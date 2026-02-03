import cv2
import pandas as pd
import os
import numpy as np
from pathlib import Path
from tqdm import tqdm

try:
    import openpyxl
    from openpyxl import load_workbook
except:
    openpyxl = None


def find_data_file(folder_path):
    """pre CSV 우선, ._ 파일 제외"""
    candidates = []
    for file in folder_path.iterdir():
        if file.name.startswith('._'):  # macOS 메타파일 제외
            continue
        if file.suffix.lower() == '.csv' and 'pre' in file.stem.lower():
            candidates.append(('csv', file))
            print(f"✅ pre CSV: {file.name}")

    if openpyxl and not candidates:
        for file in folder_path.iterdir():
            if file.name.startswith('._'):
                continue
            if file.suffix.lower() in ['.xlsx', '.xls'] and 'pre' in file.stem.lower():
                candidates.append(('excel', file))
                print(f"✅ pre 엑셀: {file.name}")

    print(f"🔍 데이터 파일 후보: {len(candidates)}개")
    return candidates[0] if candidates else None


def find_trial_file_enhanced(folder_path):
    """🔥 가장 적합한 trial 파일 적극적 탐색"""
    videos = find_videos(folder_path)
    if not videos:
        return None

    # 1. 시간형식 CSV들 우선 (2026-01-14 14-20-47.csv 패턴)
    time_csvs = list(folder_path.glob("*.csv"))
    time_csvs = [f for f in time_csvs if any(x in f.stem for x in ['202', '14-', '13-', '15-', '16-'])]

    for csv_file in sorted(time_csvs, key=lambda x: x.stat().st_size, reverse=True):
        if csv_file.name.startswith('._'):
            continue
        size_kb = csv_file.stat().st_size / 1024
        print(f"🔍 시간형식 CSV: {csv_file.name} ({size_kb:.1f}KB)")
        return ('csv', csv_file)

    # 2. 가장 큰 CSV 파일
    all_csvs = [f for f in folder_path.glob('*.csv') if not f.name.startswith('._')]
    if all_csvs:
        largest_csv = max(all_csvs, key=lambda x: x.stat().st_size)
        print(f"✅ 가장 큰 CSV: {largest_csv.name}")
        return ('csv', largest_csv)

    # 3. 엑셀 파일들
    if openpyxl:
        excel_files = [f for f in folder_path.glob('*') if f.suffix.lower() in ['.xlsx', '.xls']]
        if excel_files:
            largest_excel = max(excel_files, key=lambda x: x.stat().st_size)
            print(f"✅ 가장 큰 엑셀: {largest_excel.name}")
            return ('excel', largest_excel)

    return None


def read_trial_file(file_info):
    """🔥 파일 크기 조건 완화 + 강제 로드"""
    file_type, file_path = file_info
    print(f"🔍 trial: {file_path.name}")

    if file_path.name.startswith('._'):
        print("❌ macOS 메타파일")
        return None

    size_bytes = file_path.stat().st_size
    size_kb = size_bytes / 1024
    print(f"   📏 {size_bytes}bytes ({size_kb:.1f}KB)")

    # 🔥 100bytes 이상만 처리 (매우 완화)
    if size_bytes < 100:
        print("❌ 완전 빈 파일")
        return None

    # 엑셀 처리
    if file_type == 'excel' and openpyxl:
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            data = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2)]
            wb.close()
            df = pd.DataFrame(data)
            if len(df) > 0:
                print(f"✅ 엑셀: {len(df)}행")
                return df
        except Exception as e:
            print(f"⚠️  엑셀 실패: {e}")

    # 🔥 CSV 모든 인코딩 + header=None
    strategies = [
        {'encoding': 'utf-8-sig', 'sep': ',', 'header': None},
        {'encoding': 'utf-8', 'sep': ',', 'header': None},
        {'encoding': 'cp949', 'sep': ',', 'header': None},
        {'encoding': 'shift-jis', 'sep': ',', 'header': None},
        {'encoding': 'utf-8-sig', 'sep': '\t', 'header': None},
        {'encoding': 'utf-8', 'sep': ' ', 'header': None},
        {'encoding': 'utf-8', 'sep': None, 'header': None, 'engine': 'python'},
    ]

    for strategy in strategies:
        try:
            df = pd.read_csv(file_path, nrows=10, **strategy)
            if len(df) > 0:
                print(f"✅ 로드 성공! {len(df)}행 x {len(df.columns)}컬럼")

                # 숫자 데이터 확인
                has_numbers = False
                for i in range(min(3, len(df))):
                    try:
                        if pd.to_numeric(df.iloc[i, 1], errors='coerce') is not pd.NA:
                            has_numbers = True
                            break
                    except:
                        continue

                if has_numbers:
                    print("📋 데이터 미리보기:")
                    print(df.head(3).to_string())
                    return df
                else:
                    print("⚠️  숫자 데이터 없음")
            else:
                print(f"⚠️  빈 데이터 ({strategy['encoding']})")
        except Exception as e:
            print(f"⚠️  {strategy.get('encoding', 'unknown')} 실패")
            continue

    # 최후의 수단: 텍스트 직접 확인
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read(200)
        print(f"📄 실제 내용: {repr(content)[:100]}...")
    except:
        pass

    return None


def find_videos(folder_path):
    """비디오 찾기 (._ 제외)"""
    video_exts = ['.mp4', '.mkv', '.avi', '.mov', '.wmv', '.flv', '.webm', '.m4v']
    videos = [f for f in folder_path.iterdir() if f.suffix.lower() in video_exts and not f.name.startswith('._')]
    print(f"🎥 비디오: {len(videos)}개")
    return videos


def read_data_file(file_info):
    """Frame 데이터 읽기"""
    file_type, file_path = file_info
    if file_path.name.startswith('._'):
        return None

    if file_type == 'csv':
        return pd.read_csv(file_path)
    elif file_type == 'excel' and openpyxl:
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(row)
            wb.close()
            return pd.DataFrame(data[1:], columns=data[0])
        except:
            return None
    return None


def process_frame_data(frame_path_info, video_path, trial_df):
    """메인 처리"""
    print("🚀 처리 시작!")

    df_frame = read_data_file(frame_path_info)
    if df_frame is None or len(df_frame.columns) < 3:
        print("❌ Frame 데이터 문제")
        return False

    print(f"📊 Frame: {df_frame.shape}")
    col_frame_id, col_x, col_y = df_frame.columns[:3]

    # FPS
    cap = cv2.VideoCapture(str(video_path))
    fps = cap.get(cv2.CAP_PROP_FPS)
    cap.release()
    print(f"FPS: {fps}")

    # 시간 계산
    df_frame['F'] = df_frame[col_frame_id].apply(
        lambda x: round(int(float(x)) / fps, 3) if pd.notna(x) and str(x).strip() != '' else None
    )
    df_frame['G'] = ""

    # Trial 처리
    stats_results = []
    for i in range(min(5, len(trial_df))):
        try:
            start_time = int(float(trial_df.iloc[i, 1]))  # B열
            end_time = int(float(trial_df.iloc[i, 2])) if len(trial_df.columns) > 2 else start_time + 10

            mask_start = df_frame['F'].notna() & df_frame['F'].astype(float).apply(lambda x: int(x) == start_time)
            start_rows = df_frame[mask_start].index
            mask_end = df_frame['F'].notna() & df_frame['F'].astype(float).apply(lambda x: int(x) == end_time)
            end_rows = df_frame[mask_end].index

            if len(start_rows) > 0 and len(end_rows) > 0:
                start_row, end_row = start_rows[0], end_rows[0]
                df_frame.loc[start_row, 'G'] = f"Trial{i + 1} start"
                df_frame.loc[end_row, 'G'] = f"Trial{i + 1} end"

                trial_data = df_frame.iloc[start_row:end_row + 1]
                x_data = trial_data[col_x].dropna()
                y_data = trial_data[col_y].dropna()

                if len(x_data) > 0:
                    stats_results.append({
                        'trial': i + 1,
                        'x_cor_aver': round(x_data.mean(), 1),
                        'y_cor_aver': round(y_data.mean(), 1) if len(y_data) > 0 else 0,
                        'x_sd': round(x_data.std(), 1),
                        'y_sd': round(y_data.std(), 1) if len(y_data) > 0 else 0,
                        'row_count': len(trial_data)
                    })
                    print(f"✅ Trial{i + 1}: {start_time}s~{end_time}s")
        except Exception as e:
            print(f"⚠️  Trial{i + 1} 에러: {e}")
            continue

    # 저장
    frame_path = frame_path_info[1]
    stats_file = frame_path.parent / (frame_path.stem + "_trial_stats.csv")
    final_file = frame_path.parent / (frame_path.stem + "_final.csv")

    if stats_results:
        pd.DataFrame(stats_results).to_csv(stats_file, index=False, encoding='utf-8-sig')
        print(f"📈 통계 저장: {stats_file}")

    df_frame.to_csv(final_file, index=False, encoding='utf-8-sig')
    print(f"✅ 완료: {final_file}")
    return True


def process_folder(folder_path):
    """폴더 처리 (로그 최소화)"""
    folder = Path(folder_path)

    # Frame 파일
    frame_path_info = find_data_file(folder)
    if not frame_path_info:
        return False

    # 비디오 3개
    videos = find_videos(folder)
    if len(videos) < 3:
        return False

    video_path = min(videos, key=lambda x: x.stat().st_size)

    # Trial 파일 (강화된 탐색)
    trial_info = find_trial_file_enhanced(folder)
    if not trial_info:
        return False

    trial_df = read_trial_file(trial_info)
    if trial_df is None or len(trial_df) == 0:
        return False

    print(f"🎯 {folder.name} 처리 시작!")
    return process_frame_data(frame_path_info, video_path, trial_df)


def batch_process(root_folder):
    """배치 처리"""
    root = Path(root_folder)
    folders = [f for f in root.rglob('*') if f.is_dir()]

    print(f"📁 총 폴더: {len(folders)}개")
    success = 0

    for folder in tqdm(folders, desc="처리중"):
        try:
            if process_folder(folder):
                success += 1
        except Exception as e:
            print(f"❌ {folder.name}: {e}")

    print(f"\n🎉 완료! 성공: {success}/{len(folders)}")


if __name__ == "__main__":
    ROOT_FOLDER = r'/Volumes/ボリューム/2025_gaze_experiment'
    batch_process(ROOT_FOLDER)
